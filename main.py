"""
Agente de IA para análisis de Excel con Mistral Large
Utiliza LangChain, LangGraph, pandas, openpyxl y Streamlit
"""

import streamlit as st
import pandas as pd
import io
import json
import re
from typing import TypedDict, Annotated, List, Dict, Any
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from langchain_mistralai import ChatMistralAI
from langchain_core.messages import HumanMessage, SystemMessage, AIMessage
from langchain_core.prompts import ChatPromptTemplate
from langgraph.graph import StateGraph, END
from langgraph.prebuilt import ToolNode
from langgraph.graph.message import add_messages
import os

# ============================================================================
# CONFIGURACIÓN Y ESTADO DEL AGENTE
# ============================================================================

class AgentState(TypedDict):
    """Estado del agente que se pasa entre nodos"""
    messages: Annotated[List, add_messages]
    excel_data: Dict[str, pd.DataFrame]
    user_instruction: str
    analysis_result: Dict[str, Any]
    excel_commands: List[str]
    charts_needed: List[str]
    processed_data: Dict[str, pd.DataFrame]
    error: str


# ============================================================================
# FUNCIONES AUXILIARES PARA PROCESAMIENTO DE EXCEL
# ============================================================================

def read_excel_file(uploaded_file) -> Dict[str, pd.DataFrame]:
    """
    Lee un archivo Excel y retorna un diccionario con todas las hojas
    """
    try:
        excel_file = pd.ExcelFile(uploaded_file)
        sheets_dict = {}
        for sheet_name in excel_file.sheet_names:
            sheets_dict[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
        return sheets_dict
    except Exception as e:
        raise Exception(f"Error al leer el archivo Excel: {str(e)}")


def detect_chart_types(instruction: str) -> List[str]:
    """
    Detecta qué tipos de gráficos se necesitan basado en la instrucción
    """
    instruction_lower = instruction.lower()
    charts = []
    
    if any(word in instruction_lower for word in ['gráfico', 'grafico', 'gráfica', 'grafica', 'chart', 'visualizar', 'visualización']):
        if any(word in instruction_lower for word in ['barra', 'barras', 'bar', 'comparar', 'comparación']):
            charts.append('bar')
        if any(word in instruction_lower for word in ['pastel', 'pie', 'circular', 'proporción', 'porcentaje', 'distribución']):
            charts.append('pie')
        if any(word in instruction_lower for word in ['línea', 'linea', 'line', 'tendencia', 'temporal', 'tiempo', 'evolución']):
            charts.append('line')
        
        # Si se menciona gráfico pero no se especifica tipo, usar barras por defecto
        if not charts:
            charts.append('bar')
    
    return charts


def create_excel_with_charts(data_dict: Dict[str, pd.DataFrame], 
                             analysis_df: pd.DataFrame,
                             commands: List[str],
                             charts: List[str],
                             analysis_text: str) -> io.BytesIO:
    """
    Crea un archivo Excel con datos, análisis, comandos y gráficos
    """
    output = io.BytesIO()
    wb = Workbook()
    
    # Remover la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # 1. Agregar datos originales
    for sheet_name, df in data_dict.items():
        ws = wb.create_sheet(title=f"Original_{sheet_name}"[:31])
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
    
    # 2. Agregar hoja de análisis
    ws_analysis = wb.create_sheet(title="Análisis")
    
    # Escribir resumen del análisis
    ws_analysis['A1'] = "RESUMEN DEL ANÁLISIS"
    ws_analysis['A1'].font = Font(bold=True, size=14)
    ws_analysis['A2'] = analysis_text
    ws_analysis.merge_cells('A2:E2')
    ws_analysis['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    
    # Escribir datos del análisis
    start_row = 5
    for r_idx, row in enumerate(dataframe_to_rows(analysis_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws_analysis.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == start_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
    
    # 3. Agregar gráficos si se requieren
    chart_row = start_row + len(analysis_df) + 3
    
    if 'bar' in charts and len(analysis_df) > 0:
        chart = BarChart()
        chart.title = "Gráfico de Barras - Análisis"
        chart.y_axis.title = 'Valores'
        chart.x_axis.title = 'Categorías'
        
        data = Reference(ws_analysis, min_col=2, min_row=start_row, 
                        max_row=start_row + len(analysis_df), max_col=len(analysis_df.columns))
        cats = Reference(ws_analysis, min_col=1, min_row=start_row + 1, 
                        max_row=start_row + len(analysis_df))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        
        ws_analysis.add_chart(chart, f"A{chart_row}")
        chart_row += 18
    
    if 'pie' in charts and len(analysis_df) > 0:
        chart = PieChart()
        chart.title = "Gráfico de Pastel - Distribución"
        
        data = Reference(ws_analysis, min_col=2, min_row=start_row, 
                        max_row=start_row + len(analysis_df))
        cats = Reference(ws_analysis, min_col=1, min_row=start_row + 1, 
                        max_row=start_row + len(analysis_df))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 15
        
        ws_analysis.add_chart(chart, f"H{chart_row - 18 if 'bar' in charts else chart_row}")
        if 'bar' not in charts:
            chart_row += 18
    
    if 'line' in charts and len(analysis_df) > 0:
        chart = LineChart()
        chart.title = "Gráfico de Línea - Tendencia"
        chart.y_axis.title = 'Valores'
        chart.x_axis.title = 'Categorías'
        
        data = Reference(ws_analysis, min_col=2, min_row=start_row, 
                        max_row=start_row + len(analysis_df), max_col=len(analysis_df.columns))
        cats = Reference(ws_analysis, min_col=1, min_row=start_row + 1, 
                        max_row=start_row + len(analysis_df))
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 10
        chart.width = 20
        
        ws_analysis.add_chart(chart, f"A{chart_row}")
    
    # 4. Agregar hoja de instrucciones con comandos de Excel
    ws_commands = wb.create_sheet(title="Instrucciones")
    ws_commands['A1'] = "COMANDOS Y FÓRMULAS DE EXCEL"
    ws_commands['A1'].font = Font(bold=True, size=14)
    ws_commands.merge_cells('A1:D1')
    
    for idx, command in enumerate(commands, start=3):
        ws_commands[f'A{idx}'] = f"• {command}"
        ws_commands[f'A{idx}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws_commands.row_dimensions[idx].height = 30
    
    ws_commands.column_dimensions['A'].width = 80
    
    wb.save(output)
    output.seek(0)
    return output


# ============================================================================
# NODOS DEL AGENTE LANGGRAPH
# ============================================================================

def analyze_instruction_node(state: AgentState) -> AgentState:
    """
    Nodo que analiza la instrucción del usuario y extrae intención
    """
    try:
        llm = ChatMistralAI(
            model="mistral-large-latest",
            temperature=0.1,
            api_key=os.getenv("MISTRAL_API_KEY", "MISTRAL_API_KEY_PLACEHOLDER")
        )
        
        # Obtener información sobre los datos
        data_info = []
        for sheet_name, df in state['excel_data'].items():
            data_info.append(f"Hoja '{sheet_name}': {len(df)} filas, columnas: {', '.join(df.columns.tolist())}")
        
        data_summary = "\\n".join(data_info)
        
        system_prompt = """Eres un asistente experto en análisis de datos de Excel. 
        Tu tarea es analizar la instrucción del usuario y determinar:
        1. Qué cálculos necesita hacer (suma, promedio, máximo, mínimo, agrupaciones, etc.)
        2. Qué columnas están involucradas
        3. Qué tipo de análisis se requiere
        4. Si se necesitan gráficos
        
        Responde ÚNICAMENTE con un objeto JSON con esta estructura:
        {
            "tipo_analisis": "sumatoria|promedio|agrupacion|filtro|comparacion",
            "columnas_involucradas": ["col1", "col2"],
            "agrupacion_por": "columna_grupo o null",
            "operaciones": ["suma", "promedio", "max", "min"],
            "necesita_graficos": true|false,
            "descripcion_analisis": "descripción breve del análisis"
        }
        """
        
        user_prompt = f"""Datos disponibles:
{data_summary}

Instrucción del usuario:
{state['user_instruction']}

Analiza y responde con el JSON solicitado."""
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt)
        ]
        
        response = llm.invoke(messages)
        
        # Extraer JSON de la respuesta
        response_text = response.content
        json_match = re.search(r'\\{.*\\}', response_text, re.DOTALL)
        if json_match:
            analysis_plan = json.loads(json_match.group())
        else:
            analysis_plan = {
                "tipo_analisis": "sumatoria",
                "columnas_involucradas": [],
                "agrupacion_por": None,
                "operaciones": ["suma"],
                "necesita_graficos": False,
                "descripcion_analisis": "Análisis general de los datos"
            }
        
        state['analysis_result'] = analysis_plan
        state['charts_needed'] = detect_chart_types(state['user_instruction'])
        
        return state
        
    except Exception as e:
        state['error'] = f"Error en análisis de instrucción: {str(e)}"
        return state


def process_data_node(state: AgentState) -> AgentState:
    """
    Nodo que procesa los datos según el análisis
    """
    try:
        analysis_plan = state['analysis_result']
        excel_data = state['excel_data']
        
        # Tomar la primera hoja por defecto
        df = list(excel_data.values())[0].copy()
        
        # Determinar columnas numéricas
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        
        # Realizar análisis según el plan
        tipo = analysis_plan.get('tipo_analisis', 'sumatoria')
        agrupacion = analysis_plan.get('agrupacion_por')
        
        if agrupacion and agrupacion in df.columns:
            # Análisis agrupado
            group_col = agrupacion
            agg_dict = {}
            
            for col in numeric_cols:
                agg_dict[col] = ['sum', 'mean', 'max', 'min']
            
            result_df = df.groupby(group_col).agg(agg_dict).reset_index()
            result_df.columns = ['_'.join(col).strip('_') for col in result_df.columns.values]
            
        else:
            # Análisis simple de totales
            totals = {}
            totals['Métrica'] = ['Total', 'Promedio', 'Máximo', 'Mínimo']
            
            for col in numeric_cols:
                totals[col] = [
                    df[col].sum(),
                    df[col].mean(),
                    df[col].max(),
                    df[col].min()
                ]
            
            result_df = pd.DataFrame(totals)
        
        state['processed_data'] = {'resultado': result_df}
        
        return state
        
    except Exception as e:
        state['error'] = f"Error en procesamiento de datos: {str(e)}"
        return state


def generate_excel_commands_node(state: AgentState) -> AgentState:
    """
    Nodo que genera comandos de Excel equivalentes
    """
    try:
        llm = ChatMistralAI(
            model="mistral-large-latest",
            temperature=0.2,
            api_key=os.getenv("MISTRAL_API_KEY", "MISTRAL_API_KEY_PLACEHOLDER")
        )
        
        analysis_plan = state['analysis_result']
        
        system_prompt = """Eres un experto en Excel. Genera comandos y fórmulas de Excel para replicar el análisis.
        Proporciona instrucciones paso a paso con fórmulas específicas.
        Usa el formato: =FUNCION(RANGO) con ejemplos concretos.
        """
        
        user_prompt = f"""Análisis realizado:
{json.dumps(analysis_plan, indent=2)}

Genera 5-7 comandos/fórmulas de Excel específicas para replicar este análisis.
Lista cada comando en una línea separada."""
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt)
        ]
        
        response = llm.invoke(messages)
        
        commands = [cmd.strip() for cmd in response.content.split('\\n') if cmd.strip() and not cmd.strip().startswith('#')]
        
        # Asegurar al menos algunos comandos básicos
        if not commands:
            commands = [
                "Para sumar una columna: =SUMA(B2:B100)",
                "Para promedio: =PROMEDIO(B2:B100)",
                "Para máximo: =MAX(B2:B100)",
                "Para mínimo: =MIN(B2:B100)",
                "Para contar valores: =CONTAR(B2:B100)"
            ]
        
        state['excel_commands'] = commands
        
        return state
        
    except Exception as e:
        state['error'] = f"Error generando comandos: {str(e)}"
        state['excel_commands'] = [
            "=SUMA(rango) - Suma valores en un rango",
            "=PROMEDIO(rango) - Calcula el promedio",
            "=MAX(rango) - Encuentra el valor máximo",
            "=MIN(rango) - Encuentra el valor mínimo"
        ]
        return state


def generate_summary_node(state: AgentState) -> AgentState:
    """
    Nodo que genera resumen textual del análisis
    """
    try:
        llm = ChatMistralAI(
            model="mistral-large-latest",
            temperature=0.3,
            api_key=os.getenv("MISTRAL_API_KEY", "MISTRAL_API_KEY_PLACEHOLDER")
        )
        
        result_df = state['processed_data'].get('resultado')
        analysis_plan = state['analysis_result']
        
        system_prompt = """Eres un analista de datos. Genera un resumen ejecutivo del análisis realizado.
        Sé conciso pero informativo. Menciona los hallazgos principales."""
        
        # Convertir DataFrame a string para el prompt
        df_string = result_df.head(10).to_string() if result_df is not None else "No hay datos procesados"
        
        user_prompt = f"""Análisis realizado:
{analysis_plan.get('descripcion_analisis', 'Análisis de datos')}

Resultados (primeras filas):
{df_string}

Genera un resumen ejecutivo en 2-3 párrafos."""
        
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt)
        ]
        
        response = llm.invoke(messages)
        summary_text = response.content
        
        # Agregar el resumen al estado
        state['analysis_result']['summary'] = summary_text
        
        return state
        
    except Exception as e:
        state['error'] = f"Error generando resumen: {str(e)}"
        state['analysis_result']['summary'] = "Análisis completado. Revisa el archivo Excel para más detalles."
        return state


def check_error_node(state: AgentState) -> str:
    """
    Nodo condicional que verifica si hay errores
    """
    if state.get('error'):
        return "error"
    return "continue"


# ============================================================================
# CONSTRUCCIÓN DEL GRAFO CON LANGGRAPH
# ============================================================================

def create_agent_graph():
    """
    Crea el grafo del agente con LangGraph
    """
    workflow = StateGraph(AgentState)
    
    # Agregar nodos
    workflow.add_node("analyze_instruction", analyze_instruction_node)
    workflow.add_node("process_data", process_data_node)
    workflow.add_node("generate_commands", generate_excel_commands_node)
    workflow.add_node("generate_summary", generate_summary_node)
    
    # Definir flujo
    workflow.set_entry_point("analyze_instruction")
    
    workflow.add_edge("analyze_instruction", "process_data")
    workflow.add_edge("process_data", "generate_commands")
    workflow.add_edge("generate_commands", "generate_summary")
    workflow.add_edge("generate_summary", END)
    
    return workflow.compile()


# ============================================================================
# INTERFAZ DE STREAMLIT
# ============================================================================

def main():
    st.set_page_config(
        page_title="Agente de Análisis Excel con IA",
        page_icon="",
        layout="wide"
    )
    
    st.title("Demo Agente de Análisis Excel con AI")
    st.markdown("""
    <div style='text-align: center; color: gray;'>
        <p>Powered by Mistral Large, LangChain & LangGraph | Desarrollado con Streamlit</p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    Este agente utiliza **Mistral Large** para analizar tus datos de Excel y generar reportes automáticos.
    
    **Funcionalidades:**
    - Análisis inteligente de datos
    - Generación de reportes con sumatorias, promedios, máximos y mínimos
    - Creación automática de gráficos (barras, pastel, línea)
    - Comandos de Excel para replicar el análisis manualmente
    """)
    
    # Sidebar para configuración
    with st.sidebar:
        st.header("⚙️ Configuración")
        api_key = st.text_input(
            "Mistral API Key",
            type="password",
            value=os.getenv("MISTRAL_API_KEY", ""),
            help="Ingresa tu API key de Mistral"
        )
        
        if api_key:
            os.environ["MISTRAL_API_KEY"] = api_key
        
        st.markdown("---")
        st.markdown("### 📝 Ejemplos de instrucciones:")
        st.markdown("""
        - "Genera un reporte con la sumatoria total de ventas por estado"
        - "Calcula el promedio de ventas por mes con gráfico de línea"
        - "Muestra la distribución de ventas por estado en un gráfico de pastel"
        - "Compara las ventas entre estados con gráfico de barras"
        """)
    
    # Área principal
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("📤 Subir archivo Excel")
        uploaded_file = st.file_uploader(
            "Selecciona un archivo Excel (.xlsx)",
            type=['xlsx'],
            help="Sube tu archivo con datos para analizar"
        )
        
        if uploaded_file:
            st.success(f"✅ Archivo cargado: {uploaded_file.name}")
            
            try:
                sheets = read_excel_file(uploaded_file)
                st.info(f"📋 Hojas encontradas: {', '.join(sheets.keys())}")
                
                # Preview de datos
                with st.expander("👁️ Vista previa de datos"):
                    for sheet_name, df in sheets.items():
                        st.markdown(f"**Hoja: {sheet_name}**")
                        st.dataframe(df.head(10), use_container_width=True)
            except Exception as e:
                st.error(f"❌ Error al leer el archivo: {str(e)}")
    
    with col2:
        st.subheader("💬 Instrucción de análisis")
        user_instruction = st.text_area(
            "¿Qué análisis deseas realizar?",
            height=150,
            placeholder="Ejemplo: Genera un reporte con la sumatoria total de ventas por estado y por mes, con gráficos de barras",
            help="Describe claramente qué análisis o reporte necesitas"
        )
        
        process_button = st.button(
            "🚀 Procesar análisis",
            type="primary",
            use_container_width=True,
            disabled=not (uploaded_file and user_instruction and api_key)
        )
    
    # Procesamiento
    if process_button:
        if not api_key:
            st.error("❌ Por favor ingresa tu Mistral API Key en la barra lateral")
            return
        
        try:
            with st.spinner("🔄 Procesando tu solicitud..."):
                # Leer Excel
                excel_data = read_excel_file(uploaded_file)
                
                # Crear estado inicial
                initial_state = {
                    "messages": [],
                    "excel_data": excel_data,
                    "user_instruction": user_instruction,
                    "analysis_result": {},
                    "excel_commands": [],
                    "charts_needed": [],
                    "processed_data": {},
                    "error": ""
                }
                
                # Ejecutar agente
                progress_bar = st.progress(0)
                st.info("📊 Analizando instrucción...")
                progress_bar.progress(25)
                
                agent = create_agent_graph()
                final_state = agent.invoke(initial_state)
                
                progress_bar.progress(50)
                st.info("🔍 Procesando datos...")
                progress_bar.progress(75)
                
                st.info("📝 Generando reporte...")
                progress_bar.progress(100)
                
                # Verificar errores
                if final_state.get('error'):
                    st.error(f"❌ Error: {final_state['error']}")
                    return
                
                # Mostrar resultados
                st.success("✅ ¡Análisis completado!")
                
                # Resumen del análisis
                st.markdown("---")
                st.subheader("📋 Resumen del Análisis")
                summary = final_state['analysis_result'].get('summary', 'Análisis completado exitosamente')
                st.markdown(summary)
                
                # Vista previa de datos procesados
                st.markdown("---")
                st.subheader("📊 Datos Procesados")
                result_df = final_state['processed_data'].get('resultado')
                if result_df is not None:
                    st.dataframe(result_df, use_container_width=True)
                
                # Comandos de Excel
                st.markdown("---")
                st.subheader("🔧 Comandos de Excel")
                with st.expander("Ver comandos y fórmulas"):
                    for cmd in final_state['excel_commands']:
                        st.markdown(f"- {cmd}")
                
                # Generar archivo Excel
                st.markdown("---")
                st.subheader("💾 Descargar Reporte")
                
                excel_output = create_excel_with_charts(
                    data_dict=excel_data,
                    analysis_df=result_df,
                    commands=final_state['excel_commands'],
                    charts=final_state['charts_needed'],
                    analysis_text=summary
                )
                
                st.download_button(
                    label="📥 Descargar Excel con Análisis y Gráficos",
                    data=excel_output,
                    file_name="reporte_analisis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.balloons()
                
        except Exception as e:
            st.error(f"❌ Error al procesar: {str(e)}")
            st.exception(e)
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style='text-align: center; color: gray;'>
        <p>Powered by Mistral Large, LangChain & LangGraph | Desarrollado con Streamlit</p>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()